import openpyxl
import logging

logger = logging.getLogger(__name__)

def read_bom_from_excel(file_path, available_colors, available_sizes, plm_code=None):
    """
    Excel'den BOM verilerini okur. 
    Set siparişler için 'BOM_{plm_code}' sayfasını, tekli siparişler için ana veri sayfasını arar.
    """
    try:
        # data_only=True: Formül sonuçlarını okumak için şart
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # --- 1. Sayfa Bulma Mantığı ---
        target_sheet_name = f"BOM_{plm_code}" if plm_code else None
        sheet = None

        if target_sheet_name in workbook.sheetnames:
            sheet = workbook[target_sheet_name]
            logger.info(f"BOM verileri hedeflenen sayfadan okunuyor: {target_sheet_name}")
        else:
            # Hedef sayfa yoksa (Tekli sipariş veya eski yapı), dışlama listesini kullan
            excluded_sheets = ["Talimatlar", "İş Planı", "Lookup_Ops", "Lookup_Operations", "Sheet", "Varyant_Degerleri",
                               "Instructions", "Work Plan", "Variant_Values"]
            for name in workbook.sheetnames:
                if name not in excluded_sheets and not name.startswith("İş Planı_") and not name.startswith("Work Plan_"):
                    sheet = workbook[name]
                    logger.info(f"BOM verileri otomatik bulunan sayfadan okunuyor: {name}")
                    break
        
        if sheet is None:
            logger.error(f"Excel'de BOM verisi için uygun sayfa bulunamadı. (Hedef: {target_sheet_name})")
            return None

        # --- 2. Başlıkları ve Sütun Haritasını Oku ---
        headers = [cell.value for cell in sheet[1]]
        if not headers:
             return None
             
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        HEADER_MAP = {
            "OPERASYON": ["OPERASYON", "OPERATION"],
            "MALZEME RENGİ FARKLI MI?": ["MALZEME RENGİ FARKLI MI?", "IS MATERIAL COLOR DIFFERENT?"],
            "MALZEME KODU": ["MALZEME KODU", "MATERIAL CODE"],
            "KALEM TIPI": ["KALEM TIPI", "ITEM TYPE"],
            "MİKTAR": ["MİKTAR", "QUANTITY"],
            "BİLEŞEN ISKARTASI": ["BİLEŞEN ISKARTASI", "COMPONENT SCRAP"],
            "GENEL_RENK_SEÇİMİ": ["GENEL_RENK_SEÇİMİ", "GENERAL_COLOR_SELECTION"],
            "GENEL_BEDEN_SEÇİMİ": ["GENEL_BEDEN_SEÇİMİ", "GENERAL_SIZE_SELECTION"]
        }

        norm_col_map = {}
        for key, variants in HEADER_MAP.items():
            for variant in variants:
                if variant in col_map:
                    norm_col_map[key] = col_map[variant]
                    break

        required_keys = ["OPERASYON", "MALZEME RENGİ FARKLI MI?", "MALZEME KODU", "KALEM TIPI", "MİKTAR", "BİLEŞEN ISKARTASI", 
                         "GENEL_RENK_SEÇİMİ", "GENEL_BEDEN_SEÇİMİ"]
        
        for rk in required_keys:
            if rk not in norm_col_map:
                logger.error(f"'{sheet.title}' sayfasında '{rk}' (veya İngilizce karşılığı) sütunu bulunamadı.")
                return None

        # --- 3. Verileri Oku ---
        bom_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue

            op_val = row[norm_col_map["OPERASYON"]]
            mat_code = row[norm_col_map["MALZEME KODU"]]
            
            if not op_val or not mat_code: continue
            
            # Varyantlı mı kontrolü (X ise True, değilse False)
            is_variant_val = row[norm_col_map["MALZEME RENGİ FARKLI MI?"]]
            is_variant = True if str(is_variant_val).upper() == "X" else False

            item = {
                'OPERASYON': str(op_val).strip(),
                'MALZEME_RENGI_FARKLI_MI': is_variant,
                'MALZEME_KODU': str(mat_code).strip(),
                'KALEM_TIPI': str(row[norm_col_map["KALEM TIPI"]]).strip() if row[norm_col_map["KALEM TIPI"]] else "L",
                'MİKTAR': row[norm_col_map["MİKTAR"]],
                'BİLEŞEN_ISKARTASI': row[norm_col_map["BİLEŞEN ISKARTASI"]] or 0
            }

            # Renk Seçimi
            selected_colors = []
            genel_renk_secimi = row[norm_col_map["GENEL_RENK_SEÇİMİ"]]
            if str(genel_renk_secimi).upper() in ["TÜMÜ", "ALL"]:
                selected_colors = available_colors
            else:
                for color in available_colors:
                    col_header_tr = f"RENK_{color.upper()}"
                    col_header_en = f"COLOR_{color.upper()}"
                    idx = col_map.get(col_header_tr, col_map.get(col_header_en))
                    if idx is not None and str(row[idx]).upper() == "X":
                        selected_colors.append(color)
            item['SEÇİLİ_RENKLER'] = selected_colors

            # Beden Seçimi
            selected_sizes = []
            genel_beden_secimi = row[norm_col_map["GENEL_BEDEN_SEÇİMİ"]]
            if str(genel_beden_secimi).upper() in ["TÜMÜ", "ALL"]:
                selected_sizes = available_sizes
            else:
                for size in available_sizes:
                    col_header_tr = f"BEDEN_{size.upper()}"
                    col_header_en = f"SIZE_{size.upper()}"
                    idx = col_map.get(col_header_tr, col_map.get(col_header_en))
                    if idx is not None and str(row[idx]).upper() == "X":
                        selected_sizes.append(size)
            item['SEÇİLİ_BEDENLER'] = selected_sizes
            
            if item['SEÇİLİ_RENKLER'] and item['SEÇİLİ_BEDENLER']:
                bom_data.append(item)

        return bom_data

    except Exception as e:
        logger.exception(f"BOM okuma hatası: {e}")
        return None
    
def read_work_plan_from_excel(file_path, plm_code=None):
    """
    Set siparişler için 'İş Planı_{plm_code}' veya 'Work Plan_{plm_code}' sayfasını, 
    Tekli siparişler için 'İş Planı' veya 'Work Plan' sayfasını okur.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Arayabileceğimiz alternatif sayfa isimleri:
        possible_names = []
        if plm_code:
            possible_names.append(f"Work Plan_{plm_code}")
            possible_names.append(f"İş Planı_{plm_code}")
        possible_names.append("Work Plan")
        possible_names.append("İş Planı")
        
        target_ws = None
        for name in possible_names:
            if name in wb.sheetnames:
                target_ws = name
                break
                
        if not target_ws:
            logger.error("Excel içerisinde uygun iş planı sayfası bulunamadı.")
            return []

        ws = wb[target_ws]
        work_plan = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            step = row[0]
            operation = row[1]
            if step is not None and operation is not None:
                work_plan.append({
                    "step": step,
                    "operation": str(operation).strip()
                })
        
        logger.info(f"'{target_ws}' sayfasından {len(work_plan)} adet adım okundu.")
        return work_plan
    except Exception as e:
        logger.error(f"İş planı okuma hatası: {e}")
        return []
    
def read_variant_values_from_excel(file_path):
    """
    'Varyant_Degerleri' veya 'Variant_Values' sayfasındaki verileri ana renk bazlı gruplayarak okur.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        target_sheet = None
        for name in ["Variant_Values", "Varyant_Degerleri"]:
            if name in workbook.sheetnames:
                target_sheet = name
                break
                
        if not target_sheet:
            logger.error("Excel'de 'Variant_Values' veya 'Varyant_Degerleri' sayfası bulunamadı.")
            return None

        ws = workbook[target_sheet]
        
        # 1. Başlıkları Oku (Ana Renkleri Tespit Et)
        headers = [cell.value for cell in ws[1]]
        main_colors = []
        for h in headers[1:]:
            if h:
                h_str = str(h)
                if "ANARENK-" in h_str:
                    color_code = h_str.replace("ANARENK-", "").strip()
                    main_colors.append(color_code)
                elif "MAINCOLOR-" in h_str:
                    color_code = h_str.replace("MAINCOLOR-", "").strip()
                    main_colors.append(color_code)

        # Sonuç sözlüğünü hazırla
        color_grouped_data = {color: {} for color in main_colors}

        # 2. Satırları Oku
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            label = row[0]
            if not label: continue

            # Eğer TOPLAM satırı ise, her rengin altına TOTAL_PIECES olarak ekle
            lbl_upper = str(label).upper()
            if "TOPLAM PARÇA SAYISI" in lbl_upper or "TOTAL COMPONENT COUNT" in lbl_upper:
                for idx, color_code in enumerate(main_colors, 1):
                    total_val = row[idx]
                    color_grouped_data[color_code]["TOTAL_PIECES"] = int(total_val) if total_val else 2
                continue

            # Normal Bileşen Satırı
            for idx, color_code in enumerate(main_colors, 1):
                cell_value = row[idx]
                if cell_value is not None and str(cell_value).strip() != "":
                    try:
                        count = int(cell_value)
                    except:
                        count = 1 # Sayı değilse (X gibi) 1 kabul et
                    
                    color_grouped_data[color_code][label] = count

        logger.info(f"Varyant değerleri {len(main_colors)} ana renk için başarıyla okundu.")
        return color_grouped_data
    except Exception as e:
        logger.error(f"Varyant değerleri okuma hatası: {e}")
        return None