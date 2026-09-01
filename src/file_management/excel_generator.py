import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook.defined_name import DefinedName
from src.util.config_manager import ConfigManager
from src.util.localizer import _, get_language, get_ktsch_code_to_name
import os
import logging
import subprocess 
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

# --- SABİTLER ---
KTSCH_MAP = {
    "1000001": "Harici Dikim", "1000002": "Harici Dış Kesim", "1000003": "Harici Ütü Paket",
    "1000005": "Harici Baskı", "1000006": "Harici Biye Kesim", "1000007": "Harici Yıkama",
    "1000008": "Harici Nakış", "1000009": "Harici Çıtçıt", "1000010": "Harici İlik Düğme",
    "1000011": "Harici El İşçiliği", "1000012": "Harici Special Dikiş",
    "1000016": "APLIKE KESİM", "1000017": "POPLİN KESİM", "1000018": "ANA BEDEN KESİM",
    "1000019": "KOL UCU KESİM", "1000020": "YAKA KESİM", "1000021": "suprem kesim",
    "1000024": "oxford kesim", "1000027": "ribana kesim", "1000049": "ön beden kesim",
    "1000055": "pike kesim", "1000067": "pat kesim", "1000073": "alt pat kesim",
    "1000090": "kaskorse kesim", "1000104": "ic yaka kesim", "1000132": "kesim",
    "1000135": "dantel kesim", "1000183": "kemer kesim", "1000185": "astar kesim",
    "1000193": "garni kesim", "1000197": "ceplik kesim", "1000237": "YIKAMA",
    "1000241": "ilik düğme", "1000267": "aplike nakış", "1000271": "ÖN BASKI",
    "1000509": "tela kesim", "1000517": "tül kesim", "1000790": "cep kesim",
    "1001055": "pano kesim", "1001105": "pano baskı", "1001110": "Ara dikim",
    "1001381": "Dahili Kesim", "1001382": "JUT TEMİZLEME", "1001383": "Dahili Biye Kesim",
    "1001385": "Dahili Dikim", "1001386": "dahili yikama", "1001387": "Dahili UKP",
    "1001388": "Örgü", "1001389": "Rosso", "1001390": "Formahane", "1001391": "Jardon",
    "1001392": "Overlok", "1001393": "parça boya", "1001394": "elyaf dolum",
    "ZQM001": "Aksesuar Depo Kalite isyeri", "ZQM002": "Aksesuar Satınalma Kalite İş",
    "ZQM003": "Dokuma Depo Kontrolleri", "ZQM004": "Dokuma Tedarik Kontrolleri",
    "ZQM005": "Örme Depo Kontrolleri", "ZQM006": "Örme Tedarik Kontrolleri",
    "ZQM007": "Tübaş Kalite Kontrol"
}

KTSCH_CODE_TO_NAME = {code: name for code, name in KTSCH_MAP.items()}
ALL_OPERATION_NAMES = sorted(list(KTSCH_MAP.values()))

# --- STILLER ---
FIXED_HEADER_FILL = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid") # SteelBlue
COLOR_HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # LightBlue
SIZE_HEADER_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # DarkOrange
DYN_SIZE_HEADER_FILL = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid") # PeachPuff
GENEL_COLOR_DATA_FILL = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
GENEL_SIZE_DATA_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
VARIANT_HEADER_FILL = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid") # YellowGreen
TOTAL_ROW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")     # Yellow
def _create_instructions_sheet(workbook):
    title = "Instructions" if get_language() == "EN" else "Talimatlar"
    ws = workbook.create_sheet(title=title, index=0)
    text = _("EXCEL_INSTRUCTIONS")
    ws.merge_cells('A1:L20')
    cell = ws['A1']
    cell.value = text
    cell.font = Font(size=11); cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 100
    ws.sheet_view.showGridLines = False

def _setup_work_plan_sheet(ws, has_print, lookup_range_name):
    """İş planı sayfasını yapılandırır."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 65
    headers = ["Seq", "Operation"] if get_language() == "EN" else ["Sıra", "Operasyon"]
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=12); cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    ktsch_code_to_name = get_ktsch_code_to_name()
    steps = [
        {"Sıra": 10, "Op": ktsch_code_to_name.get("1000002")},
        {"Sıra": 20, "Op": ktsch_code_to_name.get("1000001")},
        {"Sıra": 99, "Op": ktsch_code_to_name.get("1000003")}
    ]
    if has_print:
        steps.insert(1, {"Sıra": 20, "Op": ktsch_code_to_name.get("1000005")})
        steps[2]["Sıra"] = 30

    for r, s in enumerate(steps, 2):
        ws.cell(row=r, column=1, value=s["Sıra"]).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=s["Op"]).border = THIN_BORDER
    
    dv = DataValidation(type="list", formula1=lookup_range_name, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B2:B50")

def _setup_bom_sheet(ws, available_colors, available_sizes, plan_range_name):
    """BOM sayfasını yapılandırır (Varyantlı Malzeme sütunu eklendi)."""
    
    # 1. Başlık Listesi (VARYANTLI MI? 2. sıraya eklendi)
    fixed = [
        _("EXCEL_COL_OPERATION"), 
        _("EXCEL_COL_IS_VAR_COLOR"),   # Yeni eklenen sütun
        _("EXCEL_COL_MAT_CODE"), 
        _("EXCEL_COL_ITEM_TYPE"), 
        _("EXCEL_COL_QTY"), 
        _("EXCEL_COL_COMP_SCRAP"), 
        _("EXCEL_COL_GEN_COLOR")
    ]
    
    color_prefix = "COLOR_" if get_language() == "EN" else "RENK_"
    size_prefix = "SIZE_" if get_language() == "EN" else "BEDEN_"
    genel_beden_header = "GENERAL_SIZE_SELECTION" if get_language() == "EN" else "GENEL_BEDEN_SEÇİMİ"
    
    dyn_c = [f"{color_prefix}{c.upper()}" for c in available_colors]
    dyn_s = [f"{size_prefix}{s.upper()}" for s in available_sizes]
    all_h = fixed + dyn_c + [genel_beden_header] + dyn_s
    ws.append(all_h)

    # İndeksleri dinamik olarak bulalım
    is_var_color_idx = all_h.index(_("EXCEL_COL_IS_VAR_COLOR")) + 1
    genel_renk_idx = all_h.index(_("EXCEL_COL_GEN_COLOR")) + 1
    genel_beden_idx = all_h.index(genel_beden_header) + 1

    # Başlık Stilleri
    for c_idx, h in enumerate(all_h, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = Font(bold=True); cell.border = THIN_BORDER; cell.alignment = Alignment(horizontal='center')
        
        if c_idx <= 7: cell.fill = FIXED_HEADER_FILL
        elif color_prefix in h: cell.fill = COLOR_HEADER_FILL
        elif h == genel_beden_header: cell.fill = SIZE_HEADER_FILL
        else: cell.fill = DYN_SIZE_HEADER_FILL
        
        ws.column_dimensions[get_column_letter(c_idx)].width = 50 if c_idx == 1 else max(len(h) + 5, 12)

    # Dropdownlar
    dv_op = DataValidation(type="list", formula1=plan_range_name, allow_blank=True)
    ws.add_data_validation(dv_op)
    dv_op.add("A2:A100")

    kalem_tipi_idx = all_h.index(_("EXCEL_COL_ITEM_TYPE")) + 1
    kalem_tipi_col = get_column_letter(kalem_tipi_idx)
    dv_type = DataValidation(type="list", formula1='"A,L"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f"{kalem_tipi_col}2:{kalem_tipi_col}100")

    formula_tumu = '"ALL"' if get_language() == "EN" else '"TÜMÜ"'
    dv_tumu = DataValidation(type="list", formula1=formula_tumu, allow_blank=True)
    ws.add_data_validation(dv_tumu)
    dv_tumu.add(f"{get_column_letter(genel_renk_idx)}2:{get_column_letter(genel_renk_idx)}100")
    dv_tumu.add(f"{get_column_letter(genel_beden_idx)}2:{get_column_letter(genel_beden_idx)}100")

    dv_x = DataValidation(type="list", formula1='"X"', allow_blank=True)
    ws.add_data_validation(dv_x)

    for r in range(2, 101):
        for c in range(1, len(all_h) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c == genel_renk_idx: cell.fill = GENEL_COLOR_DATA_FILL
            if c == genel_beden_idx: cell.fill = GENEL_SIZE_DATA_FILL
            if c == is_var_color_idx or color_prefix in all_h[c-1] or size_prefix in all_h[c-1]:
                dv_x.add(cell)

def _setup_variant_value_sheet(workbook, childrens, order_color_code, api_results):
    """Varyant_Degerleri sayfasını oluşturur."""
    title = "Variant_Values" if get_language() == "EN" else "Varyant_Degerleri"
    ws = workbook.create_sheet(title=title, index=1)
    
    # Başlıklar
    headers = [_("EXCEL_VARIANT_DESC_COL")] + [f"{_('EXCEL_ANARENK_PREFIX')}{c}" for c in order_color_code]
    ws.append(headers)

    # Başlık Stili (YellowGreen)
    header_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = Font(bold=True); cell.border = THIN_BORDER; cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill
    ws.column_dimensions['A'].width = 45

    # Satırları ve Eşleşmeleri Oluştur
    row_idx = 2
    for child in childrens:
        p_def = child.get("productDefiniton", "Ürün" if get_language() != "EN" else "Product")
        p_plm = child.get("plm_code", "")
        for c_color in child.get("componentColor", []):
            label = f"{p_def}-{p_plm}-{c_color}"
            ws.cell(row=row_idx, column=1, value=label).border = THIN_BORDER
            
            # Her bir sütun (Ana Renk) için API verisiyle karşılaştır
            for col_idx, m_color in enumerate(order_color_code, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
                
                # API sonucunda bu ana renk altında bu bileşen rengi var mı?
                if api_results:
                    for res in api_results:
                        if res["main_color"] == m_color:
                            if any(c["componentColorDesc"] == c_color for c in res["components"]):
                                cell.value = 1
            row_idx += 1

    # Toplam Satırı (Parça Sayısı)
    total_row = row_idx
    ws.cell(row=total_row, column=1, value=_("EXCEL_TOTAL_COMP_COUNT")).font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')
    
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for c_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(c_idx)
        cell = ws.cell(row=total_row, column=c_idx)
        cell.value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
        cell.font = Font(bold=True); cell.fill = total_fill; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        
def create_bom_template_excel(plm_id: str, style_name: str, available_colors: List[str], 
                              available_sizes: List[str], has_print: bool = False) -> Optional[str]:
    
    ConfigManager.ensure_dirs()
    """TEKLİ siparişler için Excel oluşturur."""
    output_path = os.path.join(ConfigManager.OUTPUT_EXCEL_DIR, f"{style_name}_BOM_Template_{plm_id}.xlsx")
    
    workbook = openpyxl.Workbook()
    _create_instructions_sheet(workbook)
    
    ktsch_map = get_ktsch_code_to_name()
    all_operation_names = sorted(list(ktsch_map.values()))
    
    lookup_sheet = workbook.create_sheet("Lookup_Ops")
    for idx, name in enumerate(all_operation_names, 1): lookup_sheet.cell(row=idx, column=1, value=name)
    lookup_sheet.sheet_state = 'hidden'
    workbook.defined_names.add(DefinedName('TUM_OPERASYONLAR', attr_text="'Lookup_Ops'!$A$1:$A$" + str(len(all_operation_names))))

    wp_title = "Work Plan" if get_language() == "EN" else "İş Planı"
    wp_sheet = workbook.create_sheet(wp_title, index=1)
    _setup_work_plan_sheet(wp_sheet, has_print, "TUM_OPERASYONLAR")
    workbook.defined_names.add(DefinedName('SECILI_PLAN', attr_text=f"'{wp_title}'!$B$2:$B$50"))

    bom_title = plm_id
    bom_sheet = workbook.create_sheet(title=str(bom_title), index=2)
    _setup_bom_sheet(bom_sheet, available_colors, available_sizes, "SECILI_PLAN")

    if "Sheet" in workbook.sheetnames: del workbook["Sheet"]
    workbook.save(output_path)
    if os.name == 'nt': subprocess.Popen(['start', output_path], shell=True)
    return output_path

def create_set_bom_template_excel(main_plm_id: str, style_name: str, 
                                  available_sizes: List[str], childrens: List[Dict], order_color_code: List[str], api_results: List[Dict]) -> Optional[str]:
    """SET siparişler için her child'a özel sayfalı Excel oluşturur."""
    output_path = os.path.join(ConfigManager.OUTPUT_EXCEL_DIR, f"{style_name}_BOM_Template_{main_plm_id}.xlsx")
    
    workbook = openpyxl.Workbook()
    _create_instructions_sheet(workbook)
    
    # --- YENİ SAYFA: Varyant Değerleri ---
    _setup_variant_value_sheet(workbook, childrens, order_color_code, api_results)
    
    ktsch_map = get_ktsch_code_to_name()
    all_operation_names = sorted(list(ktsch_map.values()))
    
    lookup_sheet = workbook.create_sheet("Lookup_Ops")
    for idx, name in enumerate(all_operation_names, 1): lookup_sheet.cell(row=idx, column=1, value=name)
    lookup_sheet.sheet_state = 'hidden'
    workbook.defined_names.add(DefinedName('TUM_OPERASYONLAR', attr_text="'Lookup_Ops'!$A$1:$A$" + str(len(all_operation_names))))

    for child in childrens:
        c_plm = str(child.get('plm_code'))
        c_printed = child.get('isPrinted', False)
        c_colors = child.get('componentColor', [])
        
        wp_title = f"Work Plan_{c_plm}" if get_language() == "EN" else f"İş Planı_{c_plm}"
        wp_sheet = workbook.create_sheet(wp_title)
        _setup_work_plan_sheet(wp_sheet, c_printed, "TUM_OPERASYONLAR")
        
        plan_name = f"PLAN_{c_plm}"
        workbook.defined_names.add(DefinedName(plan_name, attr_text=f"'{wp_title}'!$B$2:$B$50"))

        bom_title = f"BOM_{c_plm}"
        bom_sheet = workbook.create_sheet(bom_title)
        _setup_bom_sheet(bom_sheet, c_colors, available_sizes, plan_name)

    if "Sheet" in workbook.sheetnames: del workbook["Sheet"]
    workbook.save(output_path)
    if os.name == 'nt': subprocess.Popen(['start', output_path], shell=True)
    return output_path